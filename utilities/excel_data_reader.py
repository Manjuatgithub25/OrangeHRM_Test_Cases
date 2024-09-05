import openpyxl
import pandas as pd
from configuration.config import get_config


def data_reader(sheet_name):
    data = []
    home_dir = get_config("path", 'home_dir')
    file_path = home_dir + '\\test_data\\data.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    for row in range(2, total_rows + 1):
        for column in range(1, total_columns + 1):
            data.append(sheet.cell(row, column).value)
    return data


def excel_to_dictionary(sheet_name):
    home_dir = get_config("path", 'home_dir')
    file_path = home_dir + '\\test_data\\data.xlsx'
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    data_dict = dict(zip(df['variable'], df['value']))
    data = {key.strip(): value.strip() if isinstance(value, str) else value for key, value in data_dict.items()}
    return data


print(excel_to_dictionary('my_info'))

